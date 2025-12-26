from __future__ import annotations

from datetime import datetime, time, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from app.db import SessionLocal
from app.models import Application, User

USER_ID = "u1"
XLSX_PATH = "2025.12.xlsx"  # change if needed

STATUS_MAP = {
    "no": "applied",
    "yes": "interview",
    "reject": "rejected",
    "rejected": "rejected",
}


def norm(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def parse_date_cell(v: Any) -> Optional[datetime]:
    """
    Accepts Excel date as:
    - datetime
    - date
    - string (e.g. 2025-12-04, 12/4/2025, 04.12.2025)
    Returns datetime with time=00:00 UTC.
    """
    if v is None or norm(v) == "":
        return None

    if isinstance(v, datetime):
        d = v.date()
        return datetime.combine(d, time(0, 0, 0), tzinfo=timezone.utc)

    # openpyxl may give date as datetime already; but handle date-like too
    try:
        from datetime import date as dt_date  # noqa

        if isinstance(v, dt_date) and not isinstance(v, datetime):
            return datetime.combine(v, time(0, 0, 0), tzinfo=timezone.utc)
    except Exception:
        pass

    s = norm(v)

    # Try common formats
    fmts = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d.%m.%Y",
        "%m.%d.%Y",
        "%d-%m-%Y",
        "%m-%d-%Y",
    ]
    for fmt in fmts:
        try:
            d = datetime.strptime(s, fmt).date()
            return datetime.combine(d, time(0, 0, 0), tzinfo=timezone.utc)
        except ValueError:
            continue

    raise ValueError(f"Unrecognized date format: {s!r}")


def header_index_map(header_row: list[Any]) -> dict[str, int]:
    """
    Build case-insensitive mapping from header name -> index.
    Accepts 'company' typo as given.
    """
    m: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = norm(cell).lower()
        if key:
            m[key] = i
    return m


def main():
    xlsx_file = Path(XLSX_PATH)
    if not xlsx_file.exists():
        raise FileNotFoundError(f"Excel not found: {xlsx_file.resolve()}")

    wb = load_workbook(filename=xlsx_file, data_only=True)
    ws = wb.active  # uses first sheet; change if you want wb["SheetName"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("No rows found.")
        return

    hdr = header_index_map(list(rows[0]))
    required = ["date", "company", "title", "link", "status"]
    missing = [k for k in required if k not in hdr]
    if missing:
        raise ValueError(
            f"Missing columns in header row: {missing}. Found: {list(hdr.keys())}"
        )

    db = SessionLocal()
    try:
        # ensure user exists
        if not db.get(User, USER_ID):
            db.add(User(id=USER_ID))
            db.commit()

        last_dt: Optional[datetime] = None
        inserted = 0
        updated = 0
        skipped = 0
        bad = 0

        for r in rows[1:]:
            try:
                date_cell = r[hdr["date"]]
                dt = parse_date_cell(date_cell) or last_dt
                if dt is None:
                    # if date is missing from first row(s), fallback to now
                    dt = datetime.now(timezone.utc)
                last_dt = dt

                company = norm(r[hdr["company"]])
                title = norm(r[hdr["title"]])
                url = norm(r[hdr["link"]])
                status_raw = norm(r[hdr["status"]]).lower()
                print(f"Processing row: {dt}, {company}, {title}, {url}, {status_raw}")

                if not company or not title or not url:
                    skipped += 1
                    continue

                stage = STATUS_MAP.get(status_raw, "applied")

                # de-dupe by (user_id, url)
                existing = (
                    db.query(Application)
                    .filter(Application.user_id == USER_ID, Application.url == url)
                    .first()
                )

                if existing:
                    changed = False
                    if existing.company != company:
                        existing.company = company
                        changed = True
                    if existing.role != title:
                        existing.role = title
                        changed = True
                    if existing.stage != stage:
                        existing.stage = stage
                        changed = True
                    # keep created_at from earliest date (don’t overwrite if already set)
                    if existing.created_at is None:
                        existing.created_at = dt
                        changed = True

                    if changed:
                        existing.updated_at = datetime.now(timezone.utc)
                        updated += 1
                    else:
                        skipped += 1
                    continue

                db.add(
                    Application(
                        user_id=USER_ID,
                        company=company,
                        role=title,
                        url=url,
                        stage=stage,
                        created_at=dt,
                    )
                )
                inserted += 1

            except Exception as e:
                bad += 1
                print(f"[WARN] row skipped due to error: {e}")

        db.commit()
        print(
            f"Done. inserted={inserted}, updated={updated}, skipped={skipped}, bad={bad}"
        )

    finally:
        db.close()


if __name__ == "__main__":
    main()
